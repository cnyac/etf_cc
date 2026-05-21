"""
ingest.py — 把同花顺导出的 xlsx 读成结构化 dict。

输入约定：xlsx 有两个 sheet：
  - "中午"：表头  代码 | 名称 | 今午涨幅 | 昨涨幅 | 今午成交额 | 昨成交额
  - "收盘"：表头  代码 | 名称 | 今涨幅   | 昨涨幅 | 今成交额   | 昨成交额

session_label 含"中午"时读"中午" sheet，否则读"收盘" sheet。
session_label 含"中午"时 yest_amount 自动 ÷2（半日 vs 全日可比）。

today_date / yest_date 由调用方显式传入（如 "5月13日"），不从列名提取。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


def _parse_amount(s: Any) -> float:
    """'8.01亿' / '5519.38万' / '35654.52亿' → 元（float）。"""
    if s is None or s == "":
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip()
    if s.endswith("亿"):
        return float(s[:-1]) * 1e8
    if s.endswith("万"):
        return float(s[:-1]) * 1e4
    return float(s)


def _format_amount(v: float) -> str:
    """元 → '8.01亿' / '5519.38万'，用于呈现。"""
    if v >= 1e8:
        return f"{v / 1e8:.2f}亿"
    if v >= 1e4:
        return f"{v / 1e4:.2f}万"
    return f"{v:.2f}"


def _find_col(headers: list[str], *keywords: str) -> int:
    """在 headers 里找第一个同时包含所有 keywords 的列，返回索引。

    找不到时抛 ValueError，提示可用列名。
    """
    for i, h in enumerate(headers):
        if all(kw in h for kw in keywords):
            return i
    raise ValueError(
        f"找不到同时包含 {keywords} 的列。\n"
        f"当前所有列名：{[h for h in headers if h]}"
    )


def load_snapshot(
    xlsx_path: str | Path,
    session_label: str,
    today_date: str,
    yest_date: str,
) -> dict:
    """读 xlsx，返回标准化 snapshot dict。

    session_label: 语义标签，如 '0513中午' / '0513收盘'。
                   含"中午"时读"中午" sheet，且 yest_amount ÷2。
    today_date:    今日日期字符串，如 '5月13日'，由调用方显式传入。
    yest_date:     昨日日期字符串，如 '5月12日'，由调用方显式传入。
    """
    is_noon = "中午" in session_label
    sheet_name = "中午" if is_noon else "收盘"

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"xlsx 中找不到 '{sheet_name}' sheet。\n"
            f"现有 sheet：{wb.sheetnames}\n"
            f"请确认 xlsx 文件有 '中午' 和 '收盘' 两个 sheet。"
        )
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"'{sheet_name}' sheet 为空")
    headers = [str(c) if c is not None else "" for c in rows[0]]

    # 语义列名识别（"今午涨幅" / "今涨幅" 都能匹配 ("今", "涨幅")）
    col_code      = _find_col(headers, "代码")
    col_name      = _find_col(headers, "名称")
    col_today_pct = _find_col(headers, "今", "涨幅")
    col_yest_pct  = _find_col(headers, "昨", "涨幅")
    col_today_amt = _find_col(headers, "今", "成交额")
    col_yest_amt  = _find_col(headers, "昨", "成交额")

    items = []
    for row in rows[1:]:
        if not row or row[col_name] is None:
            continue
        name = str(row[col_name]).strip()
        if not name:
            continue
        try:
            today_pct = float(row[col_today_pct])
            yest_pct  = float(row[col_yest_pct])
        except (TypeError, ValueError):
            continue
        today_amt = _parse_amount(row[col_today_amt])
        yest_amt  = _parse_amount(row[col_yest_amt])

        # 中午时段：昨日成交额是全天数据，÷2 后才与今日半天可比
        if is_noon:
            yest_amt = yest_amt / 2

        items.append({
            "code": str(row[col_code]).strip() if row[col_code] else "",
            "name": name,
            "today_pct":        today_pct,
            "yest_pct":         yest_pct,
            "today_amount":     today_amt,
            "yest_amount":      yest_amt,
            "today_amount_str": _format_amount(today_amt),
        })

    return {
        "session_label": session_label,
        "today_date":    today_date,
        "yest_date":     yest_date,
        "items":         items,
        "total":         len(items),
    }


if __name__ == "__main__":
    import json
    import sys
    if len(sys.argv) < 5:
        print("用法: python ingest.py <xlsx> <session_label> <today_date> <yest_date>")
        print("示例: python ingest.py data/raw/ETF数据.xlsx 0513中午 '5月13日' '5月12日'")
        sys.exit(1)
    snap = load_snapshot(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
    print(json.dumps(snap, ensure_ascii=False, indent=2)[:2000])
    print(f"...\n共 {snap['total']} 个品种")
